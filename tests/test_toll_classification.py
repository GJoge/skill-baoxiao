#!/usr/bin/env python3
"""Regression tests for highway toll invoice classification."""

import os
import json
import shutil
import sys
import tempfile
from unittest.mock import patch


SKILL_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(SKILL_DIR, "scripts"))

import invoice_processor as processor  # noqa: E402
from openpyxl import Workbook, load_workbook  # noqa: E402


class FakePdf:
    def __init__(self, text):
        self.pages = [type("Page", (), {"extract_text": lambda self: text})()]

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, traceback):
        return False


class FakeReader:
    def __init__(self, path):
        text = "交易单号: OLD" if "已标记" in str(path) else ""
        self.pages = [type("Page", (), {"extract_text": lambda self: text})()]


def test_refund_change_amount_prefers_tax_inclusive_total():
    refund_text = """
    电子发票（普通发票）
    项目名称 规格型号 单 位 数 量 单 价 金 额 税率/征收率 税 额
    *现代服务*退票费 次 1 501.886792 501.89 6% 30.11
    合 计 ¥501.89 ¥30.11
    价税合计（大写） 伍佰叁拾贰元整 （小写）¥532.00
    姓名:吴国溧 客票号:781-2156992625 航班号1:MU8206 日期1:2026-04-28
    """

    with patch.object(processor.pdfplumber, "open", return_value=FakePdf(refund_text)):
        amount, source = processor.extract_amount_from_pdf("unused.pdf", "tuigai")

    assert amount == 532.00
    assert source == "价税合计(小写)"


def test_prepare_invoice_workspace_creates_backup_when_missing():
    with tempfile.TemporaryDirectory() as tmpdir:
        invoice_dir = os.path.join(tmpdir, "发票")
        backup_dir = os.path.join(tmpdir, "bak")
        os.mkdir(invoice_dir)
        with open(os.path.join(invoice_dir, "original.pdf"), "w", encoding="utf-8") as f:
            f.write("original")

        assert processor.prepare_invoice_workspace(invoice_dir, backup_dir) == "backup_created"

        assert os.path.exists(os.path.join(invoice_dir, "original.pdf"))
        assert os.path.exists(os.path.join(backup_dir, "original.pdf"))


def test_prepare_invoice_workspace_restores_clean_invoice_copy_when_backup_exists():
    with tempfile.TemporaryDirectory() as tmpdir:
        invoice_dir = os.path.join(tmpdir, "发票")
        backup_dir = os.path.join(tmpdir, "bak")
        os.mkdir(invoice_dir)
        os.mkdir(backup_dir)
        with open(os.path.join(invoice_dir, "processed.pdf"), "w", encoding="utf-8") as f:
            f.write("processed")
        with open(os.path.join(backup_dir, "original.pdf"), "w", encoding="utf-8") as f:
            f.write("original")

        assert processor.prepare_invoice_workspace(invoice_dir, backup_dir) == "restored_from_backup"

        assert not os.path.exists(os.path.join(invoice_dir, "processed.pdf"))
        assert os.path.exists(os.path.join(invoice_dir, "original.pdf"))
        assert os.path.exists(os.path.join(backup_dir, "original.pdf"))


def test_prepare_invoice_workspace_errors_without_invoice_or_backup():
    with tempfile.TemporaryDirectory() as tmpdir:
        invoice_dir = os.path.join(tmpdir, "发票")
        backup_dir = os.path.join(tmpdir, "bak")

        try:
            processor.prepare_invoice_workspace(invoice_dir, backup_dir)
        except FileNotFoundError as exc:
            assert "发票目录不存在" in str(exc)
        else:
            raise AssertionError("expected FileNotFoundError")


def test_process_invoices_does_not_write_tmp_invoice_backup():
    with tempfile.TemporaryDirectory() as tmpdir:
        invoice_dir = os.path.join(tmpdir, "发票")
        os.mkdir(invoice_dir)
        with open(os.path.join(invoice_dir, "raw.pdf"), "wb") as f:
            f.write(b"%PDF-1.4\n")

        copy_calls = []

        def fake_copy2(src, dst):
            copy_calls.append((src, dst))

        with patch.object(processor, "extract_text_from_pdf", return_value="住宿 电子发票 发票号码 价税合计"), patch.object(
            processor, "extract_amount_from_pdf", return_value=(123.0, "价税合计")
        ), patch.object(processor.shutil, "copy2", side_effect=fake_copy2):
            processor.process_invoices(invoice_dir, rename_files=True, force_write=True)

        assert not any(dst.startswith("/tmp/invoice_backup") for _, dst in copy_calls)


def test_didi_trip_matches_transaction_before_summary_invoice_with_same_amount():
    with tempfile.TemporaryDirectory() as tmpdir:
        report_path = os.path.join(tmpdir, "data_report.json")
        with open(report_path, "w", encoding="utf-8") as f:
            json.dump(
                {
                    "file_details": [
                        {"filename": "滴滴电子发票B.pdf", "type": "didi_einvoice", "amount": 51.2},
                        {
                            "filename": "滴滴出行行程报销单B.pdf",
                            "type": "didi_trip",
                            "amount": 51.2,
                            "trip_amounts": [51.2],
                        },
                    ]
                },
                f,
            )

        calls = []

        def fake_add(input_pdf, output_pdf, trans_numbers, header_text=None):
            calls.append((os.path.basename(input_pdf), list(trans_numbers)))
            return True

        with patch.object(
            processor.os,
            "listdir",
            return_value=["滴滴电子发票B.pdf", "滴滴出行行程报销单B.pdf", "微信支付账单.xlsx"],
        ), patch.object(
            processor,
            "load_wechat_bill_data",
            return_value=({51.2: ["T1"]}, {51.2: "滴滴出行"}, {}),
        ), patch.object(processor, "add_transaction_numbers_to_pdf", side_effect=fake_add), patch.object(
            processor.shutil, "move"
        ):
            processor.match_and_add_transaction_numbers(tmpdir, tmpdir, report_path, force_mark=True)

        assert calls == [("滴滴出行行程报销单B.pdf", [(1, "T1")])]


def test_existing_transaction_mark_does_not_consume_bill_transaction():
    with tempfile.TemporaryDirectory() as tmpdir:
        report_path = os.path.join(tmpdir, "data_report.json")
        with open(report_path, "w", encoding="utf-8") as f:
            json.dump(
                {
                    "file_details": [
                        {"filename": "已标记.pdf", "type": "zhusu", "amount": 20.0},
                        {"filename": "未标记.pdf", "type": "zhusu", "amount": 20.0},
                    ]
                },
                f,
            )

        calls = []

        def fake_add(input_pdf, output_pdf, trans_numbers, header_text=None):
            calls.append((os.path.basename(input_pdf), list(trans_numbers)))
            return True

        import pypdf

        with patch.object(
            processor.os,
            "listdir",
            return_value=["已标记.pdf", "未标记.pdf", "微信支付账单.xlsx"],
        ), patch.object(
            processor,
            "load_wechat_bill_data",
            return_value=({20.0: ["T1"]}, {20.0: "商户"}, {}),
        ), patch.object(pypdf, "PdfReader", FakeReader), patch.object(
            processor, "add_transaction_numbers_to_pdf", side_effect=fake_add
        ), patch.object(
            processor.shutil, "move"
        ):
            processor.match_and_add_transaction_numbers(tmpdir, tmpdir, report_path)

        assert calls == [("未标记.pdf", [(1, "T1")])]


def test_toll_invoice_and_trip_are_classified_separately():
    invoice_text = """
    电子发票（普通发票）
    项目名称 *经营租赁*代收通行费
    价税合计（小写）¥10.00
    行程信息：入口：机场高速-广东三元里站；出口：机场高速-广东机场站机场。
    """
    trip_text = """
    通行费电子行程单
    入口时间 入口站 出口时间 出口站 交易金额
    2026-04-28 14:56:24 机场高速-广东三元里站
    2026-04-28 15:15:30 机场高速-广东机场站机场 10.0元
    """

    assert processor.identify_invoice_type(invoice_text, "toll-invoice.pdf", interactive=False) == "toll"
    assert processor.identify_invoice_type(trip_text, "toll-trip.pdf", interactive=False) == "toll"
    assert processor.classify_toll_subtype(invoice_text) == "toll_einvoice"
    assert processor.classify_toll_subtype(trip_text) == "toll_trip"


def test_toll_rename_targets_are_distinct_from_didi():
    assert processor.build_invoice_rename("toll_einvoice", 1) == "高速费发票A.pdf"
    assert processor.build_invoice_rename("toll_trip", 1) == "高速费行程单A.pdf"


def test_refund_change_fee_allows_amounts_up_to_2000():
    assert processor.validate_amount("tuigai", 501.89, "退票.pdf") == (True, "金额合理")
    assert processor.validate_amount("tuigai", 2000, "退票.pdf") == (True, "金额合理")
    valid, message = processor.validate_amount("tuigai", 2000.01, "退票.pdf")
    assert valid is False
    assert "高于合理范围(¥5-2000)" in message


def test_existing_invoice_types_still_classify_normally():
    flight_text = "航空运输电子客票行程单 航班号 CA1234 承运人 中国国际航空 座位等级 经济舱"
    didi_text = "滴滴出行-行程单 DIDI TRAVEL - TRIP TABLE 快车 起点 终点 金额"
    didi_with_highway_fee_text = "滴滴出行-行程单 DIDI TRAVEL - TRIP TABLE 快车 备注 高速路桥费 3.50元"
    subway_text = "城市轨道交通 地铁 乘车码 电子发票 价税合计（小写）¥6.00"

    assert processor.identify_invoice_type(flight_text, "flight.pdf", interactive=False) == "jipiao"
    assert processor.identify_invoice_type(didi_text, "didi-trip.pdf", interactive=False) == "didi"
    assert processor.identify_invoice_type(didi_with_highway_fee_text, "didi-highway.pdf", interactive=False) == "didi"
    assert processor.classify_didi_subtype(didi_text) == "didi_trip"
    assert processor.identify_invoice_type(subway_text, "subway.pdf", interactive=False) == "didi_other"


def test_flight_city_extraction_ignores_numeric_cny_token():
    flight_text = """
    航空运输电子客票行程单
    自: 深圳 宝安 T3
    至: CNY 550.46
    至: 武汉 天河 T3
    """

    with patch.object(processor.pdfplumber, "open", return_value=FakePdf(flight_text)):
        cities, _ = processor.extract_cities_from_pdf("unused.pdf", "jipiao")

    assert set(cities) == {"深圳", "武汉"}
    assert "550.46" not in cities


def test_refund_invoice_with_airline_keywords_is_classified_as_tuigai():
    refund_text = """
    电子发票（普通发票）
    项目名称 *现代服务*退票手续费
    销方名称：中国南方航空股份有限公司
    航空运输电子客票行程单 航班号 CA1234 承运人 中国南方航空 座位等级 经济舱
    ET票号：7842178493713 吴国溧 北京-武汉 2026-05-06
    价税合计（小写）¥278.00
    """

    assert processor.identify_invoice_type(refund_text, "refund.pdf", interactive=False) == "tuigai"


def test_excel_local_transport_includes_toll_amount_once():
    with tempfile.TemporaryDirectory() as tmpdir:
        workbook_path = os.path.join(tmpdir, "biaoge.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "sheet1"
        wb.save(workbook_path)

        summary = {
            "cities_str": "",
            "huoche_count": 0,
            "huoche_amount": 0,
            "jipiao_count": 0,
            "jipiao_amount": 0,
            "tuigai_count": 0,
            "tuigai_amount": 0,
            "didi_count": 1,
            "didi_einvoice_amount": 20,
            "toll_count": 2,
            "toll_amount": 10,
            "zhusu_amount": 0,
            "jipiao_earliest_date": None,
            "jipiao_latest_date": None,
        }

        processor.write_to_excel(workbook_path, "sheet1", summary)

        saved = load_workbook(workbook_path)
        sheet = saved["sheet1"]
        assert sheet["E10"].value == 3
        assert sheet["F10"].value == 30


if __name__ == "__main__":
    test_refund_change_amount_prefers_tax_inclusive_total()
    test_prepare_invoice_workspace_creates_backup_when_missing()
    test_prepare_invoice_workspace_restores_clean_invoice_copy_when_backup_exists()
    test_prepare_invoice_workspace_errors_without_invoice_or_backup()
    test_process_invoices_does_not_write_tmp_invoice_backup()
    test_didi_trip_matches_transaction_before_summary_invoice_with_same_amount()
    test_existing_transaction_mark_does_not_consume_bill_transaction()
    test_toll_invoice_and_trip_are_classified_separately()
    test_toll_rename_targets_are_distinct_from_didi()
    test_refund_change_fee_allows_amounts_up_to_2000()
    test_existing_invoice_types_still_classify_normally()
    test_excel_local_transport_includes_toll_amount_once()
